#!/usr/bin/env python
"""
render_xlsx.py — Render JSON tabular data to a styled .xlsx workbook via openpyxl.

Pure-stdlib + openpyxl. No domain logic. Consumer skills (proposal-generator,
compliance-auditor, competitive-intel, executive-briefer) hand off a JSON
envelope and get back a styled workbook.

INPUT JSON SHAPES
-----------------
1. Top-level array of objects → single sheet named --sheet (default "Sheet1"):
   [{"col_a": 1, "col_b": "x"}, ...]

2. Top-level object whose values are arrays of objects → one sheet per key
   (sheet name = key, truncated to 31 chars per Excel limit):
   {"compliance_matrix": [...], "themes": [...], "fab_chains": [...]}

   Non-array values at the top level are ignored (e.g. executive_summary_md,
   warnings — those belong in a DOCX, not a worksheet).

STYLING (automatic)
-------------------
- Bold header row, frozen top row, autofilter across the data range.
- Column widths auto-fit to content (capped at 60 chars).
- If a column named "status" exists (case-insensitive), each row is conditionally
  filled: OK → green, PARTIAL → yellow, GAP → red.
- List values are joined with ", " for cell rendering.
- dict values are JSON-encoded for cell rendering.

EXIT CODES
----------
  0   success — output path printed to stdout
  2   bad arg / input not found / input not valid JSON / no array data found
127   openpyxl not installed (with install hint to stderr)

CLI
---
  --input <path|->   JSON source file, or "-" to read from stdin (required)
  --output <path>    Output .xlsx path (required, parent dirs auto-created)
  --sheet <name>     Sheet name when input is a top-level array (default "Sheet1")
  --title <text>     Workbook title metadata (optional)

USAGE EXAMPLE (from proposal-generator)
---------------------------------------
  run_script ../renderers/scripts/render_xlsx.py
    --input  {artifacts}/proposal_draft.json
    --output {artifacts}/compliance_matrix.xlsx
    --title  "Compliance Matrix — Program X"
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

OPENPYXL_INSTALL_HINT = """\
openpyxl is required to render .xlsx workbooks. Install with one of:

    pip install openpyxl              # any platform
    uv pip install openpyxl           # if using uv

See docs/PHASE_3E_TOOLCHAIN.md for details.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(OPENPYXL_INSTALL_HINT)
    sys.exit(127)


STATUS_FILLS = {
    "OK": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PARTIAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "GAP": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
MAX_COL_WIDTH = 60
SHEET_NAME_MAX = 31  # Excel hard limit


def _cell_value(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, list):
        return ", ".join(_cell_value(x) if isinstance(x, (str, int, float, bool)) else json.dumps(x, ensure_ascii=False) for x in v)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def _collect_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Stable column ordering: first row's keys first, then any new keys
    encountered later, in insertion order."""
    seen: dict[str, None] = {}
    for row in rows:
        for k in row.keys():
            seen.setdefault(k, None)
    return list(seen.keys())


def _write_sheet(ws, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    ws.title = sheet_name[:SHEET_NAME_MAX]
    if not rows:
        ws["A1"] = f"(no rows in '{sheet_name}')"
        return

    columns = _collect_columns(rows)
    status_col_idx = next(
        (i + 1 for i, c in enumerate(columns) if c.lower() == "status"),
        None,
    )

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        status_value: str | None = None
        for col_idx, col_name in enumerate(columns, start=1):
            value = _cell_value(row.get(col_name))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            if col_idx == status_col_idx and isinstance(value, str):
                status_value = value.strip().upper()

        if status_value and status_value in STATUS_FILLS:
            fill = STATUS_FILLS[status_value]
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Freeze header, autofilter, column widths
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            v = _cell_value(row.get(col_name))
            s = str(v)
            # account for first line of wrapped text only
            first_line_len = len(s.split("\n", 1)[0])
            if first_line_len > max_len:
                max_len = first_line_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COL_WIDTH)


def _load_input(path_arg: str) -> Any:
    if path_arg == "-":
        try:
            return json.load(sys.stdin)
        except json.JSONDecodeError as exc:
            sys.stderr.write(f"Input on stdin is not valid JSON: {exc}\n")
            sys.exit(2)

    src = Path(path_arg)
    if not src.is_file():
        sys.stderr.write(f"Input JSON file not found: {src}\n")
        sys.exit(2)
    try:
        return json.loads(src.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        sys.stderr.write(f"Input file is not valid JSON ({src}): {exc}\n")
        sys.exit(2)


def _coerce_to_sheets(data: Any, default_sheet: str) -> list[tuple[str, list[dict[str, Any]]]]:
    """Return [(sheet_name, rows)] derived from the input JSON shape."""
    if isinstance(data, list):
        if data and not all(isinstance(r, dict) for r in data):
            sys.stderr.write("Top-level array must contain objects (dicts).\n")
            sys.exit(2)
        return [(default_sheet, data)]

    if isinstance(data, dict):
        sheets: list[tuple[str, list[dict[str, Any]]]] = []
        for key, value in data.items():
            if isinstance(value, list) and value and all(isinstance(r, dict) for r in value):
                sheets.append((key, value))
        if not sheets:
            sys.stderr.write(
                "No array-of-objects values found at top level. "
                "Expected either a top-level array or an object with array-valued keys "
                "(e.g. {'compliance_matrix': [...], 'themes': [...]}).\n"
            )
            sys.exit(2)
        return sheets

    sys.stderr.write(f"Unsupported input JSON top-level type: {type(data).__name__}\n")
    sys.exit(2)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Render JSON tabular data to a styled .xlsx workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", required=True, help="JSON source path, or '-' for stdin.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name when input is a top-level array.")
    parser.add_argument("--title", default=None, help="Workbook title metadata (optional).")
    args = parser.parse_args()

    data = _load_input(args.input)
    sheets = _coerce_to_sheets(data, args.sheet)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet; we'll create our own.
    default_ws = wb.active
    wb.remove(default_ws)

    if args.title:
        wb.properties.title = args.title

    for sheet_name, rows in sheets:
        ws = wb.create_sheet()
        _write_sheet(ws, sheet_name, rows)

    try:
        wb.save(out_path)
    except OSError as exc:
        sys.stderr.write(f"Failed to write workbook ({out_path}): {exc}\n")
        return 2

    print(str(out_path.resolve()))
    return 0


if __name__ == "__main__":
    sys.exit(main())
