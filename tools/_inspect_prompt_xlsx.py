"""One-off: dump structure + content of docs/Clean_Prompt_Library.xlsx for ingestion planning."""
import json
import openpyxl

wb = openpyxl.load_workbook("docs/Clean_Prompt_Library.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
out = {}
for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n=== {s}: {ws.max_row} rows x {ws.max_column} cols ===")
    headers = [c.value for c in ws[1]]
    print("HEADERS:", headers)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append({h: v for h, v in zip(headers, r)})
    out[s] = {"headers": headers, "rows": rows}
    print(f"DATA ROWS: {len(rows)}")

with open("tools/_prompt_xlsx_dump.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2, default=str)
print("\nWrote tools/_prompt_xlsx_dump.json")
